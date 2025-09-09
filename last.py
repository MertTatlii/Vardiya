import os
import shutil
import base64
from openpyxl import Workbook, load_workbook
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QVBoxLayout, QLineEdit, QPushButton,
    QListWidget, QWidget, QHBoxLayout, QListWidgetItem, QLabel,
    QMessageBox, QFrame, QScrollArea, QFileDialog, QInputDialog
)
from PyQt6.QtCore import Qt, QSize, QByteArray
from PyQt6.QtGui import QPixmap
from PyQt6.QtWidgets import QGridLayout
from PyQt6.QtWidgets import QCheckBox
from datetime import datetime

import os
# 1) Qt’nun High-DPI desteğini tamamen kapat
os.environ["QT_ENABLE_HIGHDPI_SCALING"] = "0"
# 2) Global ölçek faktörünü 1 (yani %100) olarak sabitle
os.environ["QT_SCALE_FACTOR"] = "1"
# (isteğe bağlı) font DPI değerini klasik 96’ya çekerek metin boyutunu da kontrol altına alabilirsiniz
os.environ["QT_FONT_DPI"] = "96"

from PyQt6.QtWidgets import QApplication
import sys

# Kullanıcı veritabanı
USERS = {
    "insankaynakları": {"password": "egebantik15"},
    "depo": {"password": "egebantdepo4"},
    "bakım": {"password": "egebantbakım5"},
    "üretim": {"password": "egebantüretim6"},
}



class DataManager:
    _instance = None

    def __new__(cls):
        if cls._instance is None:
            cls._instance = super().__new__(cls)
            cls._instance.temp_data = {}
            cls._instance.unsaved_changes = {}
            cls._instance.save_path = "H:/Egebant Urfa/22-ORTAK KLASÖR/İK/Vardiya Listesi/Vardiya Programı Listeleri/"
            # 1) Sheet1’ten page->isim listesi yükle
            cls._instance.source_data = cls._instance._load_source_data()
            # 2) Sheet2’den kullanıcı->accessible_pages yükle
            cls._instance.user_access = cls._instance._load_user_access()
            cls._instance.saved_pages = set()
            cls._instance.fallback_done_pages = set()
        return cls._instance


    def _load_user_access(self) -> dict[str, list[str]]:
        """a.xlsx’in 2. sayfası (‘Kullanıcılar’)’dan:
           - 1. satır = kullanıcı adları (e.g. ebru,yunus…)
           - alt satırlar = o kullanıcının erişebileceği bölümler"""
        fn = os.path.join(self.save_path, "a.xlsx")
        if not os.path.exists(fn):
            print("Kullanıcı erişim dosyası bulunamadı:", fn)
            return {}
        wb = load_workbook(fn, data_only=True)
        sh = wb["Kullanıcılar"] if "Kullanıcılar" in wb.sheetnames else wb.worksheets[1]

        # 1. satır: başlık kullanıcı adları (küçük harfe çevir)
        headers = [str(c.value).strip().lower() for c in sh[1] if c.value]
        access = {h: [] for h in headers}
        # alt satırlardaki hücreleri ekle
        for row in sh.iter_rows(min_row=2, values_only=True):
            for idx, cell in enumerate(row):
                if idx < len(headers) and cell and str(cell).strip():
                    access[headers[idx]].append(str(cell).strip())
        return access

    def get_accessible_pages(self, username: str) -> list[str]:
        """Verilen kullanıcı adının erişebileceği bölümleri döner."""
        return self.user_access.get(username.lower(), [])

    def load_vardiya_dict(self):
        """
        Vardiya_Listesi.xlsx varsa okur ve
        { sayfa_adı: { '16:00 - 00:00': [...],
                       '08:00 - 16:00': [...],
                       '00:00 - 08:00': [...] }
        } sözlüğünü döner.
        """
        path = os.path.join(self.save_path, "Vardiya_Listesi.xlsx")
        if not os.path.exists(path):
            return {}

        try:
            wb = load_workbook(path)
            vardiya_data = {}
            for sheet in wb.worksheets:
                header_cells = list(sheet[1])
                # Başlıklara sütun indeksi bul
                col_map = {
                    cell.value: idx + 1
                    for idx, cell in enumerate(header_cells)
                    if cell.value in ["16:00 - 00:00", "08:00 - 16:00", "00:00 - 08:00"]
                }
                data = {k: [] for k in col_map}
                # 2. satırdan itibaren oku
                for row in range(2, sheet.max_row + 1):
                    for title, col in col_map.items():
                        val = sheet.cell(row=row, column=col).value
                        if val and str(val).strip():
                            data[title].append(str(val).strip())
                vardiya_data[sheet.title] = data
            return vardiya_data

        except Exception as e:
            print(f"Vardiya_Listesi dosyası okunamadı: {e}")
            return {}

    def _load_source_data(self) -> dict[str, list[str]]:
        """a.xlsx’in 1. sayfası (‘İsimler’)’den:
           - 1. satır = sayfa adları (KEY)
           - altındaki hücreler = isim listesi"""
        fn = os.path.join(self.save_path, "a.xlsx")
        if not os.path.exists(fn):
            print("Kaynak dosya bulunamadı:", fn)
            return {}
        wb = load_workbook(fn, data_only=True)
        # Sayfa adı “İsimler” yoksa default olarak ilk sheet
        sh = wb["İsimler"] if "İsimler" in wb.sheetnames else wb.worksheets[0]

        data: dict[str, list[str]] = {}
        # Başlıkları (1. satır) oku
        for col in sh.iter_cols(min_row=1, max_row=1, values_only=True):
            header = col[0]
            if header:
                data[str(header).strip().upper()] = []
        # Alt satırları her sütuna doldur
        for idx, key in enumerate(data.keys(), start=1):
            for row in sh.iter_rows(min_row=2, min_col=idx, max_col=idx, values_only=True):
                val = row[0]
                if val and str(val).strip():
                    data[key].append(str(val).strip())
        return data

    def load_page_data(self, page_name):
        if page_name in self.temp_data:
            return self.temp_data[page_name]

        excel_data = self._read_excel(page_name)
        if excel_data:
            self.temp_data[page_name] = excel_data
            return excel_data

        # Kaynak verilerden ilk sütun için verileri al (sayfa adını temizleyerek eşleştir)
        cleaned_page_name = page_name.strip().upper()
        source_column = []
        if cleaned_page_name in self.source_data:
            source_column = self.source_data[cleaned_page_name]
        else:
            # Eşleşme bulunamazsa tüm başlıkları kontrol et (içerme durumuna göre)
            for header, values in self.source_data.items():
                if cleaned_page_name in header or header in cleaned_page_name:
                    source_column = values
                    break

        self.temp_data[page_name] = {
            "isim_listesi": source_column,
            "16:00 - 00:00": [],
            "08:00 - 16:00": [],
            "00:00 - 08:00": []
        }
        return self.temp_data[page_name]

    def _read_excel(self, page_name):
        filename = f"{self.save_path}{self._clean_filename(page_name)}.xlsx"
        if not os.path.exists(filename):
            return None

        try:
            workbook = load_workbook(filename)
            sheet = workbook.active

            data = {
                "isim_listesi": [],
                "16:00 - 00:00": [],
                "08:00 - 16:00": [],
                "00:00 - 08:00": []
            }

            # Önce isim listesini kaynaktan yükle
            cleaned_page_name = page_name.strip().upper()
            if cleaned_page_name in self.source_data:
                data["isim_listesi"] = self.source_data[cleaned_page_name]
            else:
                for header, values in self.source_data.items():
                    if cleaned_page_name in header or header in cleaned_page_name:
                        data["isim_listesi"] = values
                        break

            # Sonra vardiya verilerini oku
            col_mapping = {}
            for col in range(1, 4):  # 3 sütun için (vardiyalar)
                header = sheet.cell(row=1, column=col).value
                if header in data:
                    col_mapping[header] = col

            # Verileri oku
            max_row = sheet.max_row
            for row in range(2, max_row + 1):
                for header, col in col_mapping.items():
                    value = sheet.cell(row=row, column=col).value
                    if value:
                        data[header].append(str(value).strip())

            return data
        except Exception as e:
            print(f"Excel okuma hatası: {e}")
            return None

    def save_to_excel(self, page_name, data):
        try:
            os.makedirs(self.save_path, exist_ok=True)
            filename = f"{self.save_path}{self._clean_filename(page_name)}.xlsx"

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Vardiya"

            # Başlıkları yaz (İsim Listesi hariç)
            sheet.cell(row=1, column=1, value="16:00 - 00:00")
            sheet.cell(row=1, column=2, value="08:00 - 16:00")
            sheet.cell(row=1, column=3, value="00:00 - 08:00")

            # Verileri yaz
            max_rows = max(len(data["16:00 - 00:00"]),
                           len(data["08:00 - 16:00"]),
                           len(data["00:00 - 08:00"]))

            for row in range(max_rows):
                val1 = data["16:00 - 00:00"][row] if row < len(data["16:00 - 00:00"]) else ""
                val2 = data["08:00 - 16:00"][row] if row < len(data["08:00 - 16:00"]) else ""
                val3 = data["00:00 - 08:00"][row] if row < len(data["00:00 - 08:00"]) else ""

                sheet.cell(row=row + 2, column=1, value=val1)
                sheet.cell(row=row + 2, column=2, value=val2)
                sheet.cell(row=row + 2, column=3, value=val3)

            workbook.save(filename)
            self.unsaved_changes.pop(page_name, None)

            # a.xlsx dosyasını da güncelle
            self._update_source_excel(page_name, data["isim_listesi"])

            return True
        except Exception as e:
            print(f"Kayıt hatası: {e}")
            return False

    def _update_source_excel(self, page_name, names):
        source_file = os.path.join(self.save_path, "a.xlsx")
        if not os.path.exists(source_file):
            return
        try:
            wb = load_workbook(source_file)
            # Burada aktif sayfa yerine kesinlikle 'İsimler' sayfasını alıyoruz
            sheet = wb["İsimler"] if "İsimler" in wb.sheetnames else wb.worksheets[0]

            # Temizlenmiş page_name ile başlığı bul
            cleaned = page_name.strip().upper()
            target_col = None
            for col in range(1, sheet.max_column + 1):
                header = sheet.cell(row=1, column=col).value
                if header and header.strip().upper() == cleaned:
                    target_col = col
                    break

            # Eğer yoksa yeni bir başlık ekle—but boş kolonları da doldurabilmek için önce
            # önceki boş kolon varsa orayı kullanacağız
            if target_col is None:
                # 1) var olan boş kolonları kontrol et
                for col in range(1, sheet.max_column + 1):
                    if not sheet.cell(row=1, column=col).value:
                        target_col = col
                        break
                # 2) boş kolon yoksa sonrasına ekle
                if target_col is None:
                    target_col = sheet.max_column + 1
                sheet.cell(row=1, column=target_col, value=page_name)

            # Eski verileri temizle
            for row in range(2, sheet.max_row + 1):
                sheet.cell(row=row, column=target_col, value="")

            # Yeni isimleri yaz
            for i, name in enumerate(names, start=2):
                sheet.cell(row=i, column=target_col, value=name)

            wb.save(source_file)
        except Exception as e:
            print(f"Kaynak dosya güncelleme hatası: {e}")

    # def check_missing_files(self):
    #     """Eksik dosyaları kontrol eder"""
    #     missing_files = []
    #     for page in USERS["yunus"]["accessible_pages"]:
    #         filename = f"{self.save_path}{self._clean_filename(page)}.xlsx"
    #         if not os.path.exists(filename):
    #             missing_files.append(page)
    #     return missing_files

    def create_shift_report(self):
        try:
            # 1. TÜM sayfaları kontrol et
            all_pages = set(self.source_data.keys())

            missing_files = []
            for page in all_pages:
                filename = f"{self.save_path}{self._clean_filename(page)}.xlsx"
                if not os.path.exists(filename):
                    missing_files.append(page)

            # 2. Eksik dosya varsa HEMEN iptal et (sorma)
            if missing_files:
                msg = "Aşağıdaki sayfalara ait veri dosyaları bulunamadı:\n\n"
                msg += "\n".join(f"- {name}" for name in missing_files)
                msg += "\n\nLütfen tüm dosyaları oluşturup tekrar deneyin."
                return False, msg

            # 3. Kaydetme yeri seçtir
            save_path, _ = QFileDialog.getSaveFileName(
                None,
                "Vardiya Listesini Kaydet",
                os.path.join(self.save_path, "Vardiya_Listesi.xlsx"),
                "Excel Files (*.xlsx)"
            )

            if not save_path:
                return False, "Kaydetme yeri seçilmedi"

            # 4. Workbook oluştur
            workbook = Workbook()
            workbook.remove(workbook.active)

            deleted_files = []

            # 5. TÜM sayfaları işle
            for page_name in all_pages:
                filename = f"{self.save_path}{self._clean_filename(page_name)}.xlsx"
                try:
                    sheet = workbook.create_sheet(self._clean_sheet_name(page_name))
                    source_wb = load_workbook(filename)
                    source_sheet = source_wb.active

                    # Sadece vardiya sütunlarını kopyala (İsim Listesi hariç)
                    for row in source_sheet.iter_rows(min_col=1, values_only=True):
                        sheet.append(row)

                    deleted_files.append(filename)

                except Exception as e:
                    print(f"Hata: {page_name} işlenirken - {str(e)}")
                    continue

            # # 6. Kaydet ve TÜM dosyaları sil
            # workbook.save(save_path)
            #
            # # Masaüstüne kopyala
            # desktop_copy = os.path.join(self.save_path, os.path.basename(save_path))
            # if save_path != desktop_copy:
            #     shutil.copyfile(save_path, desktop_copy)

            # Önce dosyayı seçilen konuma kaydet
            workbook.save(save_path)

            # Bugünün tarihini al ve masaüstü kopya adı oluştur
            today = datetime.now().strftime("%Y-%m-%d")
            orig_name = os.path.basename(save_path)

            copy = os.path.join(self.save_path, orig_name)

            desktop_name = f"{today}_{orig_name}"
            desktop_copy = os.path.join(self.save_path, desktop_name)
            # Eğer zaten aynı değilse kopyala
            if save_path != desktop_copy:
                shutil.copyfile(save_path, desktop_copy)
                shutil.copyfile(save_path, copy)



            # Tüm dosyaları sil
            for file in deleted_files:
                try:
                    os.remove(file)
                except Exception as e:
                    print(f"Hata: {file} silinemedi - {str(e)}")

            return True, f"Başarıyla kaydedildi:\n{save_path}\n\nMasaüstü kopyası: {desktop_copy}"

        except Exception as e:
            return False, f"Kritik hata: {str(e)}"

    def _clean_filename(self, name):
        invalid_chars = '<>:"/\\|?*'
        for char in invalid_chars:
            name = name.replace(char, '')
        return name[:31]

    def _clean_sheet_name(self, name):
        invalid_chars = '[]:*?/\\'
        for char in invalid_chars:
            name = name.replace(char, '')
        return name[:31]


class LoginWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Giriş Yap")
        self.showMaximized()

        # Ana düzen: Logo üstte, başlık ve form ortada
        layout = QVBoxLayout()
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(20)

        # Logo
        logo_label = QLabel()
        # logo_pixmap = QPixmap("logo.png")
        logo_base64 = """iVBORw0KGgoAAAANSUhEUgAAAKYAAAApCAMAAACIsKJIAAACzVBMVEUAAAAA//8AgP8Aqv8AgL8AmcwAqtUAktsAn98AjuMAmeYAldUAndgAktsAmd0Aj98AluEAnNUAlNcAmdkAktsAl9wAkN4Ald8AmdYAk9gAl9kAktsAldwAmd0AlN4Al9cAk9gAltoAmN0AlN0AltgAk9kAldoAmNsAlNsAl9wAk90AktkAldoAl9sAlNsAltwAk90Ald0Al9kAlNoAltsAlNsAldwAk9wAld0AltkAlNoAltsAk9sAldwAl9wAltoAlNoAldsAk9sAldwAltwAlN0AltoAlNoAldsAltsAlNsAltwAlNwAldoAk9oAldsAltsAlNsAldwAlNwAldoAltoAldsAltsAlNsAldwAltwAltoAlNsAldsAlNsAldwAltwAldoAltoAlNsAldsAltsAldwAltwAlNoAldoAlNsAldsAltsAldwAldwAlNoAltsAldsAltsAlNsAldwAlNoAldoAldsAldsAlNsAldwAltwAldoAltsAldsAlNsAltwAldoAldsAlNsAldsAltsAldwAldwAlNoAldsAlNsAldsAltsAldwAldwAlNoAldsAltsAldsAlNsAldwAlNoAldsAltsAldsAldsAlNsAldwAltoAldsAldsAldsAlNsAldoAldsAldsAldsAltsAldwAldoAldsAldsAltsAldsAldsAldwAldoAltsAldsAldsAldwAldoAltsAldsAldsAldsAldsAlNsAldoAldsAldsAldsAldsAldsAltsAldoAldsAldsAldsAldsAldwAldsAldsAldsAltsAldsAldsAldwAldsAlNsAldsAldsAldsAldsAldwAldsAldsAldsAldsAldsAldsAlNsAldsAldsAldsAldsAldsAldsAldsAldsAldsAldsAldsAlNsAldsAldsAldsAldsAldsAldv///8x/wZPAAAA7XRSTlMAAQIDBAUGBwgJCgwNDg8QERITFBUWFxgZGhscHR4fICEiJSYnKCkqKywtLzAxMjM0NTY3ODk6Ozw9Pj9AQUJERUZHSElKS0xNTk9QUVJTVFVWV1hZWltcXV5fYWJjZGVmZ2hpamtsbW5vcHFyc3R1d3h5ent8fX+AgYKDhIWHiIqLjI2Oj5CRkpOUlZaXmJmam52en6ChoqOkpaanqKmqrK6vsLKztLW2t7i5uru8v8DBwsPExcbHyMnKy8zNzs/Q0dLU1tfY2drb3N3e3+Dh4uPk5ebn6Onq6+zt7u/w8fLz9PX29/j5+vv8/f7pFdJoAAAAAWJLR0Tuz7fSNwAABhBJREFUWMPNmP1fFFUUhy8gggiIiAaYr1iSYZpCykqoiEoqshStokBKhqZpm2YaQkJBEJZUVKYmYon5EqCFytIqkm8ZIoq7oEDy5hvo/g/NOTN37p1lqcUP2t5f+J5z5tx9Zubec89ASI+HxvRvYxLp3TE4eePGJHti45j2Oar42IhEW8ccp4lepI7dYOuYE9XRa3Pt1j9CZsDWbkbLY8Dsuz3s6YVBq3tzytpexBytFcZrghibsWrdhr62ijkP5ipG6ebYuy/oMWH29ngimHY9nsruyWLaz8/WGdtNprb6YzmhnN9ZvbvyequxMiOQkINFwnChESfNlxX1d02mZkPRJxPNMQPSTxpbGy4eXD2Em8szqeBcdVNrW5Oh/IflA2R3Fkw7hgRk6m40t9QeSx0hug8VnYG5Gouk8ZPgU/3J1z39MLnwnJOdOfY34Y+bFIm6zmcc8eAxJ6c9pIG/l8g4C5r4jKYY6j+Bd/ZmJ43cS0D3fbNifIuQkNtK1zXpXn1ucs4NVRym5oEy43QfDjOfCzxYINGMv6vM6JzOYy7v5FKCLWM6nhdV+1WDdMcfiTPslPyXDcJvtDcyTI8GMdJ6pe6WqJZxmMK4X1fVLKpq6QZ2gfHwaIpWu7UEH3Y5jynktf91tlFM+dky5lz4Y4jAZee6Hg2cwB0fcul4YZ84qkrFq0XMlSDPq5xAD/oMjBM85tnZ/QTtV4jGLDGCt5MkvX8MDOUwTRfCoTT6HsbXDqV85KgY0MdHSWMEyQJ7EV0ryPMMqJl4F56iu99FDnM/yCl0+12Dx9efYTZ6SbusBqxU1K4gO5yklNNgRXKYLT6i8RS+/LGoQ8x2egnYo6n1OVgLQa0BtZv613KYV0DKJ8MhsF5gmJ/SQAZY+1C6qIUxjwYKuEeLmLk0gjMHW8S8DLZcOlLAWgkqFdRm6n+ZYTp0COq2nI/Lbi7DjKWBRLBOWiiJ3+Ce5DDlbvJ3sMIsYuJiz6RNzlGwtoA/E9QqetUEhukJ6o7cFlWAGccwI2jKq7hQqeW7NG3n3r35abEDyBcQSOYwo+lFv7HlbI55t2vLmAb+7aAW06v8GKaPhSZzBcNU0ZQZYF0Stfv3cjFtmt0Fc5YVmA+6/mgW+L9V3OdwhjncAuZqhhlIU4IZZp8y7tq2EnPMGVZg3gM7L5cfr4N/t6ICcE/TF1+6IiF3JsOcSlNCwfoDZbx0Kp3S66ukx9JTzFawB3Zd6N+Bf6n8ScIwB4Fq7rZDCqEWVrQylAew1E/CVsVj8yNh1oHt2/U3cxVbKJBhumAR7hZzDrUiwSpBWQ0yRgFjPWaiXq/XkXK5oGPhgc37Cqhk8GdTfyxXN/Gsd5ZpIGMqw1xJA+vAKmSH0HNyV9QzTKiR7WSPYg1my1O8gTvAQfLv4zB1/ClECtniQMxfaKAYrI9RGkEG0MBX1mAG4WEJKk8QRrICT+jB4nXeRrm8iKsxVWx9l/FnejrIY+L5SPzx6A/nznRpoYjbBncjqeSLuHO1NZj+IG8KT8mpCnmHduCxmv+hVvv+HrFLeRYP60uoK1LeenvrKROPiTdqurFjk1a76cc2XKlufId0Jn3Nis2/iu2VKwa2gW7QeAv9knt4kVVbyBmxDryzVi+VyHTzIij1O0t5n+4a128WmGfs4rZQCR9IEQMv0uL+UOrb/htTXDN0BAndT77yNw1jpKRtzFfnV81heh5VZlwYxGHOKGWB/bRB2cJdfXWTVZhT7rGUr/HLK6qYuW5ksy+Y+TqxFrft8CU8JnGMK+uQM2pS+vMFKdQ51SA19Ynsmy76NO3O94/8wCpMoiqT3kH9e/Qfco7jwtUJi9XhoSOUldBrWmR81GRHWvpcuc+5CXPUyzTqsBBv5hsGHazQE9s9H7FErRqunGrINHVCgvold0IGwmXS95OPlMJZLnxKXGykv0MPPkixFexDbG8MxeaeP1Pv2CAlgSrUQZ8ffibU2CIm7lg/vscttEVMPNMO4xegiwa/ed+1RcxIsRw0VtVIH913vGwR067U7KxJJDY5PI4o/u8TT2x1BGeW1raYOpuMx/Ni3P5Xkn8AvzOEd7Vd9yIAAAAASUVORK5CYII=
                        """
        logo_data = base64.b64decode(logo_base64)
        logo_pixmap = QPixmap()
        logo_pixmap.loadFromData(QByteArray(logo_data))
        if not logo_pixmap.isNull():
            logo_pixmap = logo_pixmap.scaled(200, 200, Qt.AspectRatioMode.KeepAspectRatio)
            logo_label.setPixmap(logo_pixmap)
            logo_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(logo_label)

        # Orta kısım için üst boşluk
        layout.addStretch(1)

        # Orta konteyner: Başlık ve form
        middle_container = QWidget()
        middle_layout = QVBoxLayout()
        middle_layout.setSpacing(20)
        middle_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)

        # Başlık
        title_label = QLabel("ŞANLIURFA EGEBANT")
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title_label.setStyleSheet("font-size: 50px; font-weight: bold;")
        middle_layout.addWidget(title_label)

        title_label = QLabel("Vardiya Hazırlama Programı")
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title_label.setStyleSheet("font-size: 30px; font-weight: bold;")
        middle_layout.addWidget(title_label)

        # Giriş formu
        self.username_input = QLineEdit()
        self.username_input.setPlaceholderText("Kullanıcı Adı")
        self.username_input.setFixedWidth(250)
        self.username_input.setFixedHeight(30)

        self.password_input = QLineEdit()
        self.password_input.setPlaceholderText("Parola")
        self.password_input.setEchoMode(QLineEdit.EchoMode.Password)
        self.password_input.setFixedWidth(250)
        self.password_input.setFixedHeight(30)


        self.username_input.returnPressed.connect(self.authenticate)
        self.password_input.returnPressed.connect(self.authenticate)

        self.login_button = QPushButton("Giriş Yap")
        self.login_button.setFixedWidth(250)
        self.login_button.setFixedHeight(50)
        self.login_button.setStyleSheet("""
            color: black;
            font-weight: bold;
            font-size: 15px;
            
        """)
        self.login_button.clicked.connect(self.authenticate)

        form_container = QWidget()
        form_layout = QVBoxLayout()
        form_layout.setSpacing(15)
        form_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        form_layout.addWidget(self.username_input)
        form_layout.addWidget(self.password_input)
        form_layout.addWidget(self.login_button)
        form_container.setLayout(form_layout)

        middle_layout.addWidget(form_container)
        middle_container.setLayout(middle_layout)
        layout.addWidget(middle_container)

        # Alt boşluk
        layout.addStretch(2)

        self.setLayout(layout)

    def authenticate(self):
        username = self.username_input.text().lower()
        password = self.password_input.text()

        if username in USERS and USERS[username]["password"] == password:
            self.main_window = PageSelectionWindow(username)
            self.main_window.showMaximized()
            self.close()
        else:
            QMessageBox.warning(self, "Hata", "Geçersiz kullanıcı adı veya parola!")



class PageSelectionWindow(QMainWindow):
    def __init__(self, username):
        super().__init__()
        # Veri yöneticisi ve kullanıcı
        self.data_manager = DataManager()
        self.username = username
        self.setWindowTitle(f"Hoşgeldiniz - {username.capitalize()}")
        self.showMaximized()

        # Ana widget & layout
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(30, 30, 30, 30)
        main_layout.setSpacing(20)

        # Logo
        logo_label = QLabel()
        # logo_pixmap = QPixmap("logo.png")
        logo_base64 = """iVBORw0KGgoAAAANSUhEUgAAAKYAAAApCAMAAACIsKJIAAACzVBMVEUAAAAA//8AgP8Aqv8AgL8AmcwAqtUAktsAn98AjuMAmeYAldUAndgAktsAmd0Aj98AluEAnNUAlNcAmdkAktsAl9wAkN4Ald8AmdYAk9gAl9kAktsAldwAmd0AlN4Al9cAk9gAltoAmN0AlN0AltgAk9kAldoAmNsAlNsAl9wAk90AktkAldoAl9sAlNsAltwAk90Ald0Al9kAlNoAltsAlNsAldwAk9wAld0AltkAlNoAltsAk9sAldwAl9wAltoAlNoAldsAk9sAldwAltwAlN0AltoAlNoAldsAltsAlNsAltwAlNwAldoAk9oAldsAltsAlNsAldwAlNwAldoAltoAldsAltsAlNsAldwAltwAltoAlNsAldsAlNsAldwAltwAldoAltoAlNsAldsAltsAldwAltwAlNoAldoAlNsAldsAltsAldwAldwAlNoAltsAldsAltsAlNsAldwAlNoAldoAldsAldsAlNsAldwAltwAldoAltsAldsAlNsAltwAldoAldsAlNsAldsAltsAldwAldwAlNoAldsAlNsAldsAltsAldwAldwAlNoAldsAltsAldsAlNsAldwAlNoAldsAltsAldsAldsAlNsAldwAltoAldsAldsAldsAlNsAldoAldsAldsAldsAltsAldwAldoAldsAldsAltsAldsAldsAldwAldoAltsAldsAldsAldwAldoAltsAldsAldsAldsAldsAlNsAldoAldsAldsAldsAldsAldsAltsAldoAldsAldsAldsAldsAldwAldsAldsAldsAltsAldsAldsAldwAldsAlNsAldsAldsAldsAldsAldwAldsAldsAldsAldsAldsAldsAlNsAldsAldsAldsAldsAldsAldsAldsAldsAldsAldsAldsAlNsAldsAldsAldsAldsAldsAldv///8x/wZPAAAA7XRSTlMAAQIDBAUGBwgJCgwNDg8QERITFBUWFxgZGhscHR4fICEiJSYnKCkqKywtLzAxMjM0NTY3ODk6Ozw9Pj9AQUJERUZHSElKS0xNTk9QUVJTVFVWV1hZWltcXV5fYWJjZGVmZ2hpamtsbW5vcHFyc3R1d3h5ent8fX+AgYKDhIWHiIqLjI2Oj5CRkpOUlZaXmJmam52en6ChoqOkpaanqKmqrK6vsLKztLW2t7i5uru8v8DBwsPExcbHyMnKy8zNzs/Q0dLU1tfY2drb3N3e3+Dh4uPk5ebn6Onq6+zt7u/w8fLz9PX29/j5+vv8/f7pFdJoAAAAAWJLR0Tuz7fSNwAABhBJREFUWMPNmP1fFFUUhy8gggiIiAaYr1iSYZpCykqoiEoqshStokBKhqZpm2YaQkJBEJZUVKYmYon5EqCFytIqkm8ZIoq7oEDy5hvo/g/NOTN37p1lqcUP2t5f+J5z5tx9Zubec89ASI+HxvRvYxLp3TE4eePGJHti45j2Oar42IhEW8ccp4lepI7dYOuYE9XRa3Pt1j9CZsDWbkbLY8Dsuz3s6YVBq3tzytpexBytFcZrghibsWrdhr62ijkP5ipG6ebYuy/oMWH29ngimHY9nsruyWLaz8/WGdtNprb6YzmhnN9ZvbvyequxMiOQkINFwnChESfNlxX1d02mZkPRJxPNMQPSTxpbGy4eXD2Em8szqeBcdVNrW5Oh/IflA2R3Fkw7hgRk6m40t9QeSx0hug8VnYG5Gouk8ZPgU/3J1z39MLnwnJOdOfY34Y+bFIm6zmcc8eAxJ6c9pIG/l8g4C5r4jKYY6j+Bd/ZmJ43cS0D3fbNifIuQkNtK1zXpXn1ucs4NVRym5oEy43QfDjOfCzxYINGMv6vM6JzOYy7v5FKCLWM6nhdV+1WDdMcfiTPslPyXDcJvtDcyTI8GMdJ6pe6WqJZxmMK4X1fVLKpq6QZ2gfHwaIpWu7UEH3Y5jynktf91tlFM+dky5lz4Y4jAZee6Hg2cwB0fcul4YZ84qkrFq0XMlSDPq5xAD/oMjBM85tnZ/QTtV4jGLDGCt5MkvX8MDOUwTRfCoTT6HsbXDqV85KgY0MdHSWMEyQJ7EV0ryPMMqJl4F56iu99FDnM/yCl0+12Dx9efYTZ6SbusBqxU1K4gO5yklNNgRXKYLT6i8RS+/LGoQ8x2egnYo6n1OVgLQa0BtZv613KYV0DKJ8MhsF5gmJ/SQAZY+1C6qIUxjwYKuEeLmLk0gjMHW8S8DLZcOlLAWgkqFdRm6n+ZYTp0COq2nI/Lbi7DjKWBRLBOWiiJ3+Ce5DDlbvJ3sMIsYuJiz6RNzlGwtoA/E9QqetUEhukJ6o7cFlWAGccwI2jKq7hQqeW7NG3n3r35abEDyBcQSOYwo+lFv7HlbI55t2vLmAb+7aAW06v8GKaPhSZzBcNU0ZQZYF0Stfv3cjFtmt0Fc5YVmA+6/mgW+L9V3OdwhjncAuZqhhlIU4IZZp8y7tq2EnPMGVZg3gM7L5cfr4N/t6ICcE/TF1+6IiF3JsOcSlNCwfoDZbx0Kp3S66ukx9JTzFawB3Zd6N+Bf6n8ScIwB4Fq7rZDCqEWVrQylAew1E/CVsVj8yNh1oHt2/U3cxVbKJBhumAR7hZzDrUiwSpBWQ0yRgFjPWaiXq/XkXK5oGPhgc37Cqhk8GdTfyxXN/Gsd5ZpIGMqw1xJA+vAKmSH0HNyV9QzTKiR7WSPYg1my1O8gTvAQfLv4zB1/ClECtniQMxfaKAYrI9RGkEG0MBX1mAG4WEJKk8QRrICT+jB4nXeRrm8iKsxVWx9l/FnejrIY+L5SPzx6A/nznRpoYjbBncjqeSLuHO1NZj+IG8KT8mpCnmHduCxmv+hVvv+HrFLeRYP60uoK1LeenvrKROPiTdqurFjk1a76cc2XKlufId0Jn3Nis2/iu2VKwa2gW7QeAv9knt4kVVbyBmxDryzVi+VyHTzIij1O0t5n+4a128WmGfs4rZQCR9IEQMv0uL+UOrb/htTXDN0BAndT77yNw1jpKRtzFfnV81heh5VZlwYxGHOKGWB/bRB2cJdfXWTVZhT7rGUr/HLK6qYuW5ksy+Y+TqxFrft8CU8JnGMK+uQM2pS+vMFKdQ51SA19Ynsmy76NO3O94/8wCpMoiqT3kH9e/Qfco7jwtUJi9XhoSOUldBrWmR81GRHWvpcuc+5CXPUyzTqsBBv5hsGHazQE9s9H7FErRqunGrINHVCgvold0IGwmXS95OPlMJZLnxKXGykv0MPPkixFexDbG8MxeaeP1Pv2CAlgSrUQZ8ffibU2CIm7lg/vscttEVMPNMO4xegiwa/ed+1RcxIsRw0VtVIH913vGwR067U7KxJJDY5PI4o/u8TT2x1BGeW1raYOpuMx/Ni3P5Xkn8AvzOEd7Vd9yIAAAAASUVORK5CYII=
                        """
        logo_data = base64.b64decode(logo_base64)
        logo_pixmap = QPixmap()
        logo_pixmap.loadFromData(QByteArray(logo_data))
        if not logo_pixmap.isNull():
            logo_pixmap = logo_pixmap.scaled(200, 200, Qt.AspectRatioMode.KeepAspectRatio)
            logo_label.setPixmap(logo_pixmap)
            logo_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        main_layout.addWidget(logo_label)

        # Başlık
        title = QLabel("Lütfen bir sayfa seçiniz")
        title.setStyleSheet("font-size: 24px; font-weight: bold;")
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        main_layout.addWidget(title)

        # Bölüm butonları için scroll + grid
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)

        inner_widget = QWidget()
        self.grid_layout = QGridLayout()
        self.grid_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.grid_layout.setHorizontalSpacing(20)
        self.grid_layout.setVerticalSpacing(15)

        # Butonları ekle
        self.add_buttons_to_grid()

        inner_widget.setLayout(self.grid_layout)
        scroll.setWidget(inner_widget)
        main_layout.addWidget(scroll)

        # Alt butonlar: Vardiya Raporu & Bölüm Ekle & Bölüm Sil
        footer = QWidget()
        footer_layout = QHBoxLayout(footer)
        footer_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)

        shift_btn = QPushButton("Vardiya Listesi Oluştur")
        shift_btn.setFixedWidth(300)
        shift_btn.setStyleSheet("""
            QPushButton {
                background-color: #ff9800;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 15px;
                font-size: 16px;
            }
            QPushButton:hover {
                background-color: #f57c00;
            }
        """
        )
        shift_btn.clicked.connect(self.create_shift_report)
        footer_layout.addWidget(shift_btn)

        add_sec_btn = QPushButton("Bölüm Ekle")
        add_sec_btn.setFixedWidth(300)
        add_sec_btn.setStyleSheet("""
            QPushButton {
                background-color: #29b6f6;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 15px;
                font-size: 16px;
            }
            QPushButton:hover {
                background-color: #0288d1;
            }
        """
        )
        add_sec_btn.clicked.connect(self.add_section)
        footer_layout.addWidget(add_sec_btn)

        delete_sec_btn = QPushButton("Bölüm Sil")
        delete_sec_btn.setFixedWidth(300)
        delete_sec_btn.setStyleSheet("""
            QPushButton {
                background-color: #e57373;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 15px;
                font-size: 16px;
            }
            QPushButton:hover {
                background-color: #d32f2f;
            }
        """
        )
        delete_sec_btn.clicked.connect(self.remove_section)
        footer_layout.addWidget(delete_sec_btn)

        main_layout.addWidget(footer)
        central_widget.setLayout(main_layout)

    def add_buttons_to_grid(self):
        # Önce temizle
        for i in reversed(range(self.grid_layout.count())):
            w = self.grid_layout.itemAt(i).widget()
            if w:
                w.setParent(None)

        # Excel'den gelen bölümleri al
        accessible_pages = self.data_manager.get_accessible_pages(self.username)
        row, col = 0, 0
        max_cols = 3

        for page in accessible_pages:
            btn = QPushButton(page)
            btn.setFixedSize(300, 60)
            saved = page in self.data_manager.saved_pages
            normal_bg = "#4caf50" if saved else "#5c6bc0"
            hover_bg = "#388e3c" if saved else "#3949ab"
            btn.setStyleSheet(f"""
                QPushButton {{
                    background-color: {normal_bg};
                    color: white;
                    border: none;
                    border-radius: 4px;
                    padding: 15px;
                    font-size: 16px;
                }}
                QPushButton:hover {{
                    background-color: {hover_bg};
                }}
            """
            )
            btn.clicked.connect(lambda _, p=page: self.open_page(p))
            self.grid_layout.addWidget(btn, row, col)
            col += 1
            if col >= max_cols:
                col = 0
                row += 1

    def open_page(self, page_name):
        # Kaydedilmemiş değişiklik uyarısı
        if page_name in self.data_manager.unsaved_changes:
            reply = QMessageBox.question(
                self, 'Kaydedilmemiş Veri',
                'Kaydedilmemiş verileriniz var. Devam etmek istiyor musunuz?',
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            if reply == QMessageBox.StandardButton.No:
                return

        # Sayfayı aç
        self.page_window = ContentWindow(page_name, self.username)
        self.page_window.showMaximized()
        self.hide()

    def create_shift_report(self):
        success, message = self.data_manager.create_shift_report()
        if success:
            QMessageBox.information(self, "Başarılı", message)
        else:
            QMessageBox.critical(self, "Hata", message)

    def add_section(self):
        section, ok = QInputDialog.getText(self, "Yeni Bölüm Ekle", "Bölüm adı:")
        if not (ok and section.strip()):
            return
        section = section.strip()

        dm = self.data_manager
        # Global kopya kontrolü
        if section.upper() in (name.upper() for name in dm.source_data.keys()):
            QMessageBox.warning(self, "Uyarı", f"'{section}' adlı bölüm zaten mevcut!")
            return
        # Kullanıcıya özel kontrol
        if section in dm.get_accessible_pages(self.username):
            QMessageBox.warning(self, "Uyarı", "Bu bölüme zaten erişiminiz var!")
            return

        fn = os.path.join(dm.save_path, "a.xlsx")
        wb = load_workbook(fn)

        # Kullanıcılar sheet'ine bölüm adı ekle
        sh_users = wb["Kullanıcılar"] if "Kullanıcılar" in wb.sheetnames else wb.worksheets[1]
        user_col = next(
            (cell.column for cell in sh_users[1]
             if str(cell.value).strip().lower() == self.username.lower()),
            None
        )
        if user_col is None:
            QMessageBox.warning(self, "Hata", "Kullanıcı sütunu bulunamadı!")
            return
        row = 2
        while sh_users.cell(row, user_col).value:
            row += 1
        sh_users.cell(row, user_col).value = section

        # İsimler sheet'ine ilk boş sütuna başlık olarak ekle
        sh_names = wb["İsimler"] if "İsimler" in wb.sheetnames else wb.worksheets[0]
        col_index = None
        for col in range(1, sh_names.max_column + 1):
            if not sh_names.cell(1, col).value:
                col_index = col
                break
        if col_index is None:
            col_index = sh_names.max_column + 1
        sh_names.cell(1, col_index).value = section

        wb.save(fn)

        # Önbellekleri güncelle ve UI’ı yenile
        dm.source_data = dm._load_source_data()
        dm.user_access = dm._load_user_access()
        self.add_buttons_to_grid()
        QMessageBox.information(self, "Tamam", f"'{section}' bölümü eklendi.")

    def remove_section(self):
        dm = self.data_manager
        # Silinecek bölümleri al
        pages = dm.get_accessible_pages(self.username)
        if not pages:
            QMessageBox.information(self, "Uyarı", "Silinecek bölüm yok.")
            return
        section, ok = QInputDialog.getItem(
            self, "Bölüm Sil", "Lütfen silmek istediğiniz bölümü seçin:", pages, 0, False
        )
        if not (ok and section):
            return

        reply = QMessageBox.question(
            self, "Onay", f"'{section}' bölümü kalıcı olarak silinecek. Devam edilsin mi?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        if reply != QMessageBox.StandardButton.Yes:
            return

        fn = os.path.join(dm.save_path, "a.xlsx")
        wb = load_workbook(fn)

        # Kullanıcılar sheet'inden değeri sil ve aşağı kaydır
        sh_users = wb["Kullanıcılar"] if "Kullanıcılar" in wb.sheetnames else wb.worksheets[1]
        user_col = next(
            (cell.column for cell in sh_users[1]
             if str(cell.value).strip().lower() == self.username.lower()),
            None
        )
        if user_col:
            for row in range(2, sh_users.max_row + 1):
                if str(sh_users.cell(row, user_col).value).strip() == section:
                    for r in range(row, sh_users.max_row):
                        sh_users.cell(r, user_col).value = sh_users.cell(r+1, user_col).value
                    sh_users.cell(sh_users.max_row, user_col).value = None
                    break

        # İsimler sheet'inden ilgili sütunu sil
        sh_names = wb["İsimler"] if "İsimler" in wb.sheetnames else wb.worksheets[0]
        del_col = None
        for cell in sh_names[1]:
            if str(cell.value).strip() == section:
                del_col = cell.column
                break
        if del_col:
            sh_names.delete_cols(del_col)

        wb.save(fn)

        # 4.5) Bölüme ait sayfa dosyasını sil
        # DataManager._clean_filename ile güvenli dosya adı
        file_to_delete = os.path.join(
            dm.save_path,
            f"{dm._clean_filename(section)}.xlsx"
        )
        if os.path.exists(file_to_delete):
            try:
                os.remove(file_to_delete)
            except Exception as e:
                QMessageBox.warning(
                    self, "Uyarı",
                    f"'{section}.xlsx' silinirken hata oluştu:\n{e}"
                )

        # Önbellekleri güncelle ve UI’ı yenile
        dm.source_data = dm._load_source_data()
        dm.user_access = dm._load_user_access()
        self.add_buttons_to_grid()
        QMessageBox.information(self, "Tamam", f"'{section}' bölümü silindi.")



class ContentWindow(QMainWindow):
    def __init__(self, page_name, username):
        super().__init__()
        self.page_name = page_name
        self.username = username
        self.data_manager = DataManager()
        self.setWindowTitle(f"{page_name} - {username.capitalize()}")
        self.showMaximized()

        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QVBoxLayout(central)
        main_layout.setContentsMargins(20,20,20,20)
        main_layout.setSpacing(20)

        # Title
        title = QLabel(page_name)
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title.setStyleSheet("QLabel { font-size:24px; font-weight:bold; }")
        main_layout.addWidget(title)

        # Load data
        self.page_data = self.data_manager.load_page_data(page_name)
        all_old = self.data_manager.load_vardiya_dict()
        self.old_vardiya = all_old.get(self.page_name, {})

        # --- Fallback: sadece bu sayfa ilk kez açılıyorsa ---
        dm = self.data_manager
        if self.page_name not in dm.fallback_done_pages:
            for shift in ["16:00 - 00:00", "08:00 - 16:00", "00:00 - 08:00"]:
                if not self.page_data.get(shift) and self.old_vardiya.get(shift):
                    # Kopya alarak doldur
                    self.page_data[shift] = self.old_vardiya[shift].copy()
            # Bir daha bu sayfada fallback uygulanmasın
            dm.fallback_done_pages.add(self.page_name)

        # Source list
        scroll_src = QScrollArea()
        scroll_src.setWidgetResizable(True)
        scroll_src.setFrameShape(QFrame.Shape.NoFrame)
        self.source_list_widget = SourceListWidget(
            "İsim Listesi", self.page_data["isim_listesi"], self
        )
        scroll_src.setWidget(self.source_list_widget)

        # Shifts
        self.shift_widgets = {}
        def make_shift(slot):
            scroll = QScrollArea(); scroll.setWidgetResizable(True)
            scroll.setFrameShape(QFrame.Shape.NoFrame)
            widget = ShiftListWidget(slot, self.page_data[slot], self)
            scroll.setWidget(widget)
            return scroll, widget
        scroll1, self.shift_widgets["16:00 - 00:00"] = make_shift("16:00 - 00:00")
        scroll2, self.shift_widgets["08:00 - 16:00"] = make_shift("08:00 - 16:00")
        scroll3, self.shift_widgets["00:00 - 08:00"] = make_shift("00:00 - 08:00")

        # Grid layout
        grid = QGridLayout()
        grid.setHorizontalSpacing(20); grid.setVerticalSpacing(15)
        grid.addWidget(scroll_src, 0,0)
        grid.addWidget(scroll1, 0,1)
        grid.addWidget(scroll2, 0,2)
        grid.addWidget(scroll3, 0,3)

        # Add new item
        add_ctrl = QWidget(); hbox = QHBoxLayout(add_ctrl)
        hbox.setContentsMargins(50,0,50,0); hbox.setSpacing(10)
        self.add_input = QLineEdit(); self.add_input.setFixedWidth(200)
        self.add_input.setPlaceholderText("İsim Giriniz")

        self.add_input.returnPressed.connect(self.add_new_item)

        add_btn = QPushButton("Yeni Kişi Ekle"); add_btn.setFixedWidth(100)
        add_btn.clicked.connect(self.add_new_item)
        hbox.addWidget(self.add_input); hbox.addWidget(add_btn)
        grid.addWidget(add_ctrl, 1,0)
        main_layout.addLayout(grid)

        # Save & back
        save_btn = QPushButton("KAYDET"); save_btn.setFixedWidth(200)
        save_btn.setStyleSheet("background-color: #66BB6A;")
        save_btn.clicked.connect(self.save_data)
        main_layout.addWidget(save_btn, alignment=Qt.AlignmentFlag.AlignCenter)
        back_btn = QPushButton("Geri Dön"); back_btn.setFixedWidth(200)
        back_btn.clicked.connect(self.go_back)
        main_layout.addWidget(back_btn, alignment=Qt.AlignmentFlag.AlignCenter)

    def add_new_item(self):
        name = self.add_input.text().strip()
        if name and name not in self.page_data["isim_listesi"]:
            self.page_data["isim_listesi"].append(name)
            self.data_manager.unsaved_changes[self.page_name] = True
            self.refresh_view(); self.add_input.clear()
        else:
            QMessageBox.warning(self, "Uyarı", "Bu kişi zaten listede var!")

    def bulk_move_to_shift(self, source_shift, target_shift):
        lw = (self.source_list_widget.list_widget
              if source_shift=="isim_listesi"
              else self.shift_widgets[source_shift].list_widget)
        selected = [
            lw.itemWidget(lw.item(i)).name
            for i in range(lw.count())
            if lw.itemWidget(lw.item(i)).checkbox.isChecked()
        ]
        if not selected: return
        for name in selected:
            if source_shift!="isim_listesi" and name in self.page_data[source_shift]:
                self.page_data[source_shift].remove(name)
            for s in ["16:00 - 00:00","08:00 - 16:00","00:00 - 08:00"]:
                if name in self.page_data.get(s,[]): self.page_data[s].remove(name)
            if name not in self.page_data[target_shift]:
                self.page_data[target_shift].append(name)
        self.data_manager.unsaved_changes[self.page_name]=True
        self.refresh_view()

    def handle_move(self, source_shift, target_shift, name):
        lw = (self.source_list_widget.list_widget
              if source_shift=="isim_listesi"
              else self.shift_widgets[source_shift].list_widget)
        # Check if any selected
        any_sel = any(
            lw.itemWidget(lw.item(i)).checkbox.isChecked()
            for i in range(lw.count())
        )
        if any_sel:
            self.bulk_move_to_shift(source_shift, target_shift)
        else:
            self._move_single(source_shift, target_shift, name)

    def _move_single(self, source_shift, target_shift, name):
        if source_shift!="isim_listesi" and name in self.page_data[source_shift]:
            self.page_data[source_shift].remove(name)
        for s in ["16:00 - 00:00","08:00 - 16:00","00:00 - 08:00"]:
            if name in self.page_data.get(s,[]): self.page_data[s].remove(name)
        if name not in self.page_data[target_shift]:
            self.page_data[target_shift].append(name)
        self.data_manager.unsaved_changes[self.page_name]=True
        self.refresh_view()

    def refresh_view(self):
        self.new_window = ContentWindow(self.page_name, self.username)
        self.new_window.showMaximized(); self.close()

    def save_data(self):
        if self.data_manager.save_to_excel(self.page_name, self.page_data):
            # Başarıyla kaydedilen sayfayı işaretle
            self.data_manager.saved_pages.add(self.page_name)
            QMessageBox.information(self, "Başarılı", "Veriler başarıyla kaydedildi!")
        else:
            QMessageBox.critical(self, "Hata", "Kayıt sırasında hata oluştu!")

    def delete_item(self, name, shift=None):
        if shift:
            if name in self.page_data[shift]:
                self.page_data[shift].remove(name)
                self.data_manager.unsaved_changes[self.page_name]=True; self.refresh_view()
        else:
            ans = QMessageBox.question(
                self, 'Onay', f"'{name}' silinsin mi?",
                QMessageBox.StandardButton.Yes|QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No)
            if ans==QMessageBox.StandardButton.Yes:
                if name in self.page_data["isim_listesi"]:
                    self.page_data["isim_listesi"].remove(name)
                for s in ["16:00 - 00:00","08:00 - 16:00","00:00 - 08:00"]:
                    if name in self.page_data[s]: self.page_data[s].remove(name)
                self.data_manager.unsaved_changes[self.page_name]=True; self.refresh_view()

    def go_back(self):
        if self.page_name in self.data_manager.unsaved_changes:
            ans = QMessageBox.question(
                self, 'Kaydedilmemiş Veri',
                'Kaydedilmemiş verileriniz var. Geri dönmek ister misiniz?',
                QMessageBox.StandardButton.Yes|QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No)
            if ans==QMessageBox.StandardButton.No: return
        self.prev = PageSelectionWindow(self.username)
        self.prev.show(); self.close()

class SourceListWidget(QWidget):
    def __init__(self, title, initial_data, parent):
        super().__init__()
        self.parent = parent
        self.names = initial_data.copy()

        # Layout ayarları
        self.layout = QVBoxLayout()
        self.layout.setContentsMargins(0, 0, 0, 0)
        self.layout.setSpacing(0)

        # Başlık
        title_label = QLabel(title)
        title_label.setStyleSheet(
            """
            QLabel {
                font-size: 16px;
                font-weight: bold;
                padding: 10px;
                background-color: #f0f0f0;
                border: 1px solid #d0d0d0;
                border-radius: 4px 4px 0 0;
            }
            """
        )
        self.layout.addWidget(title_label)

        # "Tümünü Seç" checkbox
        self.select_all_cb = QCheckBox("Tümünü Seç")
        self.select_all_cb.toggled.connect(self.on_select_all_toggled)
        self.layout.addWidget(self.select_all_cb)

        # Liste widget
        self.list_widget = QListWidget()
        self.list_widget.setStyleSheet(
            """
            QListWidget { border: 1px solid #d0d0d0; border-top: none; border-radius: 0 0 4px 4px; font-size: 14px; min-height: 300px; }
            QListWidget::item { height: 50px; }
            QListWidget::item:selected { background-color: #e8f4ff; }
            """
        )
        self.list_widget.itemClicked.connect(self.on_item_clicked)

        # Öğeleri ekle
        for name in self.names:
            self.create_list_item(name)

        self.layout.addWidget(self.list_widget)
        self.setLayout(self.layout)

    def create_list_item(self, name):
        item = QListWidgetItem()
        item.setSizeHint(QSize(0, 50))
        widget = EditableListItem(name, self.parent, is_source=True)
        self.list_widget.addItem(item)
        self.list_widget.setItemWidget(item, widget)

    def on_item_clicked(self, item):
        widget = self.list_widget.itemWidget(item)
        # Diğer öğelerin butonlarını gizle
        for i in range(self.list_widget.count()):
            w = self.list_widget.itemWidget(self.list_widget.item(i))
            if w and w != widget:
                w.hide_buttons()
        widget.show_buttons()

    def on_select_all_toggled(self, checked: bool):
        print(f"[SourceListWidget] select_all toggled: {checked}")
        for i in range(self.list_widget.count()):
            w = self.list_widget.itemWidget(self.list_widget.item(i))
            w.checkbox.setChecked(checked)


class ShiftListWidget(QWidget):
    def __init__(self, time_slot, initial_data, parent):
        super().__init__()
        self.time_slot = time_slot
        self.parent = parent
        self.names = initial_data.copy()

        # Layout ayarları
        self.layout = QVBoxLayout()
        self.layout.setContentsMargins(0, 0, 0, 0)
        self.layout.setSpacing(0)

        # Başlık
        title_label = QLabel(time_slot)
        title_label.setStyleSheet(
            """
            QLabel {
                font-size: 16px;
                font-weight: bold;
                padding: 10px;
                background-color: #f0f0f0;
                border: 1px solid #d0d0d0;
                border-radius: 4px 4px 0 0;
            }
            """
        )
        self.layout.addWidget(title_label)

        # "Tümünü Seç" checkbox
        self.select_all_cb = QCheckBox("Tümünü Seç")
        self.select_all_cb.toggled.connect(self.on_select_all_toggled)
        self.layout.addWidget(self.select_all_cb)

        # Liste widget
        self.list_widget = QListWidget()
        self.list_widget.setStyleSheet(
            """
            QListWidget { border: 1px solid #d0d0d0; border-top: none; border-radius: 0 0 4px 4px; font-size: 14px; min-height: 300px; }
            QListWidget::item { height: 50px; }
            QListWidget::item:selected { background-color: #e8f4ff; }
            """
        )
        self.list_widget.itemClicked.connect(self.on_item_clicked)

        # Öğeleri ekle
        for name in self.names:
            self.create_list_item(name)

        self.layout.addWidget(self.list_widget)
        self.setLayout(self.layout)

    def create_list_item(self, name):
        item = QListWidgetItem()
        item.setSizeHint(QSize(0, 50))
        widget = EditableListItem(name, self.parent, is_source=False, shift_name=self.time_slot)
        self.list_widget.addItem(item)
        self.list_widget.setItemWidget(item, widget)

    def on_item_clicked(self, item):
        widget = self.list_widget.itemWidget(item)
        for i in range(self.list_widget.count()):
            w = self.list_widget.itemWidget(self.list_widget.item(i))
            if w and w != widget:
                w.hide_buttons()
        widget.show_buttons()

    def on_select_all_toggled(self, checked: bool):
        print(f"[ShiftListWidget {self.time_slot}] select_all toggled: {checked}")
        for i in range(self.list_widget.count()):
            w = self.list_widget.itemWidget(self.list_widget.item(i))
            w.checkbox.setChecked(checked)




class EditableListItem(QWidget):
    def __init__(self, name, parent=None, is_source=False, shift_name=None):
        super().__init__(parent)
        self.parent = parent
        self.name = name
        self.is_source = is_source
        self.shift_name = "isim_listesi" if is_source else shift_name

        # Layout
        self.layout = QHBoxLayout()
        self.layout.setContentsMargins(15, 8, 15, 8)
        self.layout.setSpacing(10)

        # Checkbox
        self.checkbox = QCheckBox()
        self.layout.addWidget(self.checkbox)

        # Label
        self.label = QLabel(name)
        self.label.setStyleSheet(
            "QLabel { font-size: 14px; padding: 5px; qproperty-alignment: AlignVCenter;}"
        )
        self.layout.addWidget(self.label, stretch=1)

        if not self.is_source and hasattr(self.parent, 'old_vardiya'):
            eski = self.parent.old_vardiya  # { '16:00 - 00:00': [...], ... }
            # Bu isim eski hangi vardiyada yer alıyor?
            eski_shift = next(
                (s for s, lst in eski.items() if self.name in lst),
                None
            )
            if eski_shift:
                # Aynı vardiya ise kırmızı, farklı ise yeşil
                renk = 'red' if eski_shift == self.shift_name else 'green'
                self.label.setStyleSheet(
                    f"QLabel {{ "
                    f"font-size: 14px; "
                    f"padding: 5px; "
                    f"qproperty-alignment: AlignVCenter; "
                    f"color: {renk}; "
                    f"}}"
                )

        # Buttons
        self.add_btn1 = QPushButton("1'e Taşı")
        self.add_btn2 = QPushButton("2'ye Taşı")
        self.add_btn3 = QPushButton("3'e Taşı")
        self.delete_btn = QPushButton("Sil")
        for btn in (self.add_btn1, self.add_btn2, self.add_btn3, self.delete_btn):
            btn.setFixedHeight(30)
            btn.setStyleSheet(
                "QPushButton { background-color: #f0f0f0; border:1px solid #d0d0d0;"
                " border-radius:5px; padding:6px 12px; font-size:13px; }"
                "QPushButton:hover { background-color:#e0e0e0; }"
            )
            btn.hide()
        self.layout.addWidget(self.add_btn1)
        self.layout.addWidget(self.add_btn2)
        self.layout.addWidget(self.add_btn3)
        self.layout.addWidget(self.delete_btn)

        self.setLayout(self.layout)

        # Connections: single or bulk move
        src = self.shift_name
        self.add_btn1.clicked.connect(lambda: self.parent.handle_move(src, "16:00 - 00:00", self.name))
        self.add_btn2.clicked.connect(lambda: self.parent.handle_move(src, "08:00 - 16:00", self.name))
        self.add_btn3.clicked.connect(lambda: self.parent.handle_move(src, "00:00 - 08:00", self.name))
        self.delete_btn.clicked.connect(lambda: self.parent.delete_item(self.name, None if self.is_source else self.shift_name))

    def show_buttons(self):
        self.add_btn1.show()
        self.add_btn2.show()
        self.add_btn3.show()
        self.delete_btn.show()

    def hide_buttons(self):
        self.add_btn1.hide()
        self.add_btn2.hide()
        self.add_btn3.hide()
        self.delete_btn.hide()

if __name__ == "__main__":
    app = QApplication(sys.argv)
    login_window = LoginWindow()
    login_window.show()
    app.exec()
